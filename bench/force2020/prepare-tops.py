#!/usr/bin/env python3
"""
Convert the FORCE 2020 / NPD formations spreadsheet into the plain CSVs the
app's own importers already read — Well,Surface,MD for tops.ts::parseTopsCsv
and Well,X,Y for heads.ts::parseWellHeadsCsv. The bench harness then exercises
the real import path instead of a one-off reader written just for this data.

Source: NPD_Lithostratigraphy_member_formations_all_wells.xlsx from
https://zenodo.org/records/4351156 (CC BY 4.0, Norwegian Petroleum Directorate
via the FORCE 2020 competition). Not redistributed here — download it
yourself per bench/force2020/README.md.
"""
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit('нужен openpyxl: pip3 install openpyxl')

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / 'datasets' / 'force2020' / 'formations.xlsx'
if not SRC.exists():
    sys.exit(f'нет файла {SRC} — сначала скачайте его (см. bench/force2020/README.md)')

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(min_row=2, values_only=True))

tops_path = ROOT / 'datasets' / 'force2020' / 'tops.csv'
heads_path = ROOT / 'datasets' / 'force2020' / 'heads.csv'

heads_seen = {}
n_tops = 0
with tops_path.open('w', newline='', encoding='utf-8') as ft:
    tw = csv.writer(ft)
    tw.writerow(['Well', 'Surface', 'MD'])
    for r in rows:
        well, surface, x, y, _z, md = r[0], r[1], r[2], r[3], r[4], r[5]
        if not well or not surface or md is None:
            continue
        tw.writerow([well, surface, md])
        n_tops += 1
        # Wellhead X/Y repeats on every pick row; keep the first seen, same
        # rule the app's own Petrel-tops importer uses.
        if well not in heads_seen and x is not None and y is not None:
            heads_seen[well] = (x, y)

with heads_path.open('w', newline='', encoding='utf-8') as fh:
    hw = csv.writer(fh)
    hw.writerow(['Well', 'X', 'Y'])
    for well, (x, y) in heads_seen.items():
        hw.writerow([well, x, y])

print(f'{n_tops} разбивок, {len(heads_seen)} устьев -> {tops_path.name}, {heads_path.name}')
